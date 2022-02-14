import sys
from tkinter import CENTER
from tkinter.font import BOLD
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5 import uic
import cv2
import numpy as np
#라즈베리 파이의 경우 아래 주석 해제
import picamera
from os import listdir, makedirs
from os.path import isdir, isfile, join
from datetime import datetime
from openpyxl import load_workbook
from Detect import face_detector, run, train, face_classifier, trains
from Register import face_extractor, take_pictures
import sys
sys.path.append("Detect")
import Detect as dt


class MyApp(QWidget):
    #############################회원가입#############################
    # 얼굴을 저장할 디렉토리
    face_dirs = 'faces/'
    # haarcascade 학습모듈 로드
    face_classifier = cv2.CascadeClassifier('haarcascades/haarcascade_frontalface_default.xml')

    # 얼굴 검출 함수
    def face_extractor(img):
        # 사진을 회색으로 변환 (연산량을 줄일 수 있음)
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray,1.3,5)
        # 얼굴이 없으면 넘어감
        if faces == ():
            return None
        # 얼굴이 있으면 얼굴 부위만 이미지로 만들고
        for(x,y,w,h) in faces:
            cropped_face = img[y:y+h, x:x+w]
        # 자른 이미지를 반환
        return cropped_face

    # 사진 촬영 함수
    def take_pictures(self,name):
        # 해당 이름의 폴더가 없다면 생성
        if not isdir(self.face_dirs+name):
            makedirs(self.face_dirs+name)

        # 카메라 ON    
        cap = cv2.VideoCapture(0)
        count = 0

        while True:
            # 카메라로 부터 사진 한장 읽어 오기
            ret, frame = cap.read()
            # 사진에서 얼굴 검출 , 얼굴이 검출되었다면 
            if face_extractor(frame) is not None:
                count+=1
                # 촬영한 사진을 200 x 200 사이즈로 조정
                face = cv2.resize(face_extractor(frame),(200,200))
                # 흑백으로 바꿈
                face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
                # 200x200 흑백 사진을 faces/얼굴 이름/숫자.jpg 로 저장
                file_name_path = self.face_dirs + name + '/' + str(count) +'.jpg'
                #cv2.imwrite(file_name_path,face)

                dst = cv2.flip(face, flipCode=1)
                #이미지 파일 저장
                retval, buffer = cv2.imencode('.jpg', dst)
                if (retval):
                    with open(file_name_path, 'wb') as f_writer:
                        f_writer.write(buffer)
                
                # 촬영된 사진이 몇번째 사진인지 창 안에 표시
                cv2.putText(face,str(count),(50,50),cv2.FONT_HERSHEY_COMPLEX,1,(0,255,0),2)
                cv2.imshow('Detected Face',face)
            else:
                print("Face not Found")
                pass
            
            # 얼굴 사진 30장을 촬영했거나 enter키 누르면 종료
            if cv2.waitKey(1)==13 or count==30:
                break

        # 촬영 종료
        cap.release()
        # 창 닫음
        cv2.destroyAllWindows()
        print('학습 샘플이미지 추출 완료')

    #############################출석체크#############################

    # 얼굴 인식용 haarcascade 로딩
    face_classifier = cv2.CascadeClassifier('./haarcascades/haarcascade_frontalface_default.xml')    

    # 사용자 얼굴 학습
    def train(name):
        data_path = 'faces/' + name + '/'
        #파일만 리스트로 만듬
        face_pics = [f for f in listdir(data_path) if isfile(join(data_path,f))]
        
        Training_Data, Labels = [], []
        
        # 사용자의 사진을 불러옴
        for i, files in enumerate(face_pics):
            image_path = data_path + face_pics[i]
            n = np.fromfile(image_path, dtype=np.uint8)
            img = cv2.imdecode(n, cv2.IMREAD_GRAYSCALE)

            Training_Data.append(np.asarray(img, dtype=np.uint8))
            Labels.append(i)
        # 학습할 데이터(이미지)가 없으면 함수 종료
        if len(Labels) == 0:
            print("학습할 이미지가 없습니다.")
            return None
        # Lables 를 32비트 정수로 변환함
        Labels = np.asarray(Labels, dtype=np.int32)
        # 모델 생성
        model = cv2.face.LBPHFaceRecognizer_create()
        # 학습
        model.train(np.asarray(Training_Data), np.asarray(Labels))
        print(name + " : 모델 학습 완료")

        #학습 모델 리턴
        return model

    # 여러 사용자 학습
    def trains():
        #faces 폴더의 하위 폴더를 학습
        data_path = 'faces/'
        # 폴더만 색출
        model_dirs = [f for f in listdir(data_path) if isdir(join(data_path,f))]
        
        #학습 모델 저장할 딕셔너리
        models = {}
        # 각 폴더에 있는 얼굴들 학습
        for model in model_dirs:
            print('model :' + model)
            # 학습 시작
            result = train(model)
            # 학습이 안되었다면 패스!
            if result is None:
                continue
            # 학습되었으면 저장
            models[model] = result

        # 학습된 모델 딕셔너리 리턴
        return models   

    #얼굴 검출
    def face_detector(img, size = 0.5):
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_classifier.detectMultiScale(gray,1.3,5)
            if faces == ():
                return img,[]
            for(x,y,w,h) in faces:
                cv2.rectangle(img, (x,y),(x+w,y+h),(0,255,255),2)
                roi = img[y:y+h, x:x+w]
                roi = cv2.resize(roi, (200,200))
            return img,roi   #검출된 좌표에 사각 박스 그리고(img), 검출된 부위를 잘라(roi) 전달

    # 인식 시작
    def run(models):    
        print("출석체크 프로그램 시작")
        #카메라 열기 
        cap = cv2.VideoCapture(0)
        try:
            while True:
                #카메라로 부터 사진 한장 읽기 
                ret, frame = cap.read()
                # 얼굴 검출 시도 
                image, face = face_detector(frame)
                try:            
                    min_score = 999       #예측된 사람일 확률
                    min_score_name = ""   #예측된 사람의 이름
                
                    #검출된 사진을 흑백으로 변환 
                    face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)

                    #위에서 학습한 모델로 예측시도
                    for key, model in models.items():
                        result = model.predict(face)                
                        #result[1] 에 들어가는 값은 신뢰도 임
                        if min_score > result[1]:
                            min_score = result[1]
                            min_score_name = key
                            uid = min_score_name.split("_")[0]
                            uname = min_score_name.split("_")[1]
                    # 0에 가까울 수록 정확하다.         
                    if min_score < 500:
                        # 정확도 
                        confidence = int(100*(1-(min_score)/300))
                        # 유사도 화면에 표시 
                        display_string = str(confidence)+'% Confidence it is ' + uid
                    cv2.putText(image,display_string,(100,120), cv2.FONT_HERSHEY_COMPLEX,1,(250,120,255),2)
                    
                    #80 보다 크면 동일 인물로 간주해 UnLocked! 
                    if confidence > 80:
                        cv2.putText(image,  uid, (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
                        # 저장할 파일의 경로와 이름지정
                        now = str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        load_wb = load_workbook("attend.xlsx")
                        load_ws = load_wb['Sheet1']
                        '''
                        for cell in load_ws['C']:  
                            if cell.value == now:
                                pass
                            else:
                        '''
                        write_ws= load_wb.active
                        write_ws.append([uid,uname,now])
                        load_wb.save("attend.xlsx")
                        break      
                        
                #얼굴 검출 안됨 
                except:
                    cv2.imshow('Face Cropper', image)
                    pass
                #Enter 입력으로 프로그램 종료
                if cv2.waitKey(1)==13:
                    break
            #카메라 끄기
            cap.release()
            #열려있는 창 닫기
            cv2.destroyAllWindows()
        except KeyboardInterrupt:
            pass


    #############################메인#############################

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        txt1 = QLabel("한결컴퓨터학원 출석체크 프로그램")
        txt1.setAlignment(Qt.AlignCenter)
        txt1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        txtid = QLabel("ID:")

        self.eid = QLineEdit()

        txtname = QLabel("이름:")

        self.ename = QLineEdit()


        #회원등록 버튼
        btn1 = QPushButton('회원등록', self)
        btn1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        btn1.clicked.connect(self.OnBtn1Clicked)

        #출석체크 버튼
        btn2 = QPushButton('출석체크', self)
        btn2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        btn2.clicked.connect(self.OnBtn2Clicked)

        #메인화면
        txtbox = QHBoxLayout()
        txtbox.addWidget(txt1)
        idbox = QHBoxLayout()
        idbox.addWidget(txtid)
        idbox.addWidget(self.eid)
        namebox = QHBoxLayout()
        namebox.addWidget(txtname)
        namebox.addWidget(self.ename)
        infobox = QVBoxLayout()
        infobox.addLayout(idbox)
        infobox.addLayout(namebox)
        registerbox = QHBoxLayout()
        registerbox.addLayout(infobox)
        registerbox.addWidget(btn1)
        btnbox = QVBoxLayout()
        btnbox.addWidget(btn2)
        
        
        #레이아웃 추가
        layout = QVBoxLayout()
        layout.addLayout(txtbox)
        layout.addLayout(registerbox)
        layout.addLayout(btnbox)

        self.setLayout(layout)
        

        self.setWindowTitle('QPushButton')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def OnBtn1Clicked(self):
        id = self.eid.text()
        name = self.ename.text()
        if(id == '' or name == ''):
            QMessageBox.about(self,'Error','아이디와 이름을 모두 입력해주세요!')
        else:
            load_wb = load_workbook("user.xlsx")
            load_ws = load_wb['Sheet1']
            take_pictures(id+"_"+name)
            write_ws= load_wb.active
            write_ws.append([id,name])
            load_wb.save("user.xlsx")
            QMessageBox.about(self,'회원등록','사용자 입력 완료!')
            self.eid.setText("")
            self.ename.setText("")
        

    def OnBtn2Clicked(self):
        models = trains()
        while True:
            run(models)
            QMessageBox.about(self,'출석체크','ID:'+dt.uid+'\n이름:'+dt.uname+'\n출석체크 완료!')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
